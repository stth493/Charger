from flask import Flask, render_template, jsonify

from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

import smtplib
import requests
import xmltodict
import openpyxl
import pandas as pd
import ipaddress
import os
import json
import time
import threading
from multiprocessing.dummy import Pool as ThreadPool 
from datetime import datetime, timedelta
from bs4 import BeautifulSoup


app = Flask(__name__, template_folder='templates', static_folder='static', static_url_path='/')


api_data_cache = {}
last_update_time = None
last_alert_times = {}  # Dictionary to track last alert time for each charger


charger_url = {
        'Argus' :       { 'url'  : 'http://charger_ip/livedata.xml?0=',
                          'user' : '',
                          'pass' : '4321'
                        },
        'DongAH' :      { 'url'  : 'http://charger_ip/rect_footer.html',
                          'user' : 'admin',
                          'pass' : '4321'
                        },
        'Great power' : { 'url'  : 'http://charger_ip/rect_footer.html',
                          'user' : 'admin',
                          'pass' : '654321'
                        },
    }


def multithread_ping(ips: list):
    pool = ThreadPool(8)
    return pool.starmap(ping_device,zip(ips))

def ping_device(host: str) -> str:
    try: 
        ipaddress.ip_address(host)
        if os.system("ping -n 1 " +host) == 0:
            return {host : "On-line"}
        else:
            return {host : "Off-line"}
    except:
        return {host : "Invalid IP address"}

def update_status_charger(ping_result:list):
    status_dict = {k: v for d in ping_result for k, v in d.items()}
    df = pd.read_excel("IP Charger of SYMC.xlsx", sheet_name="IP Charger")
    df["Remark"] = df["IP Address"].apply(lambda ip: status_dict.get(ip, df["Remark"]))
    write_to_excel_append(df["Remark"].tolist(), column = 'H')

def write_to_excel_append(data: list, column: str):
    workbook = openpyxl.load_workbook('IP Charger of SYMC.xlsx')
    sheet = workbook.active
    for i, value in enumerate(data, start=2):  # start=1 for 1-based indexing
        # Convert value to string if it's not None, otherwise use empty string
        excel_value = str(value) if value is not None else ""
        sheet[f'{column}{i}'] = excel_value
    workbook.save('IP Charger of SYMC.xlsx')

def read_charger_list() -> list:
    '''
    If flag is True then return all charger list otherwise will return only online
    '''
    Charger_dict = {}
    try:
        workbook = openpyxl.load_workbook('IP Charger of SYMC.xlsx', data_only=True)
        # Get the specified sheet or the first sheet
        sheet = workbook.active
    except FileNotFoundError:
        raise FileNotFoundError(f"Excel file not found in script folder")
    except Exception as e:
        raise Exception(f"Error reading Excel file: {str(e)}")
    
    Site_col = None
    Brand_col = None
    IP_col = None
    Remark_col = None
    for col in range(1, sheet.max_column + 1):
        if sheet.cell(row=1, column=col).value == "Site Name":
            Site_col = col
        if sheet.cell(row=1, column=col).value == "Brand":
            Brand_col = col    
        if sheet.cell(row=1, column=col).value == "IP Address":
            IP_col = col
        if sheet.cell(row=1, column=col).value == "Remark":
            Remark_col = col
    if IP_col is None or Site_col is None or Brand_col is None:
        raise ValueError("Expected column not found in the Excel file")
    for row in range(2, sheet.max_row + 1):
        charger_name = sheet.cell(row=row, column=Site_col).value
        brand = sheet.cell(row=row, column=Brand_col).value
        status = sheet.cell(row=row, column=Remark_col).value
        #if brand != 'Argus' or status != 'On-line': continue
        IP = sheet.cell(row=row, column=IP_col).value
        if charger_name == None: continue
        Info_dict = {}
        Info_dict['Brand'] = brand
        Info_dict['IP'] = IP
        Info_dict['Status'] = status
        Charger_dict[charger_name] = Info_dict
    return Charger_dict

def send_email(subject:str, body:str, from_email:str, to_email:str):
    # Create the email
    msg = MIMEMultipart()
    msg['From'] = f'Charger Rectifier Monitoring'
    msg['To'] = ", ".join(to_email)
    msg['Subject'] = subject

    msg.attach(MIMEText(body, 'plain'))

    try:
        # Connect to the server and send the email
        server = smtplib.SMTP('10.10.10.160', 25)
        server.starttls()
        server.login('sitthi.s@symphony.net.th', 'Jorjaa101227!')
        server.sendmail(from_email, to_email, msg.as_string())
        server.quit()
        print("E-mail sent successfully")
    except smtplib.SMTPAuthenticationError as e:
        print(f"Authentication failed: {e}")
    except smtplib.SMTPConnectError as e:
        print(f"Failed to connect to the server: {e}")
    except smtplib.SMTPHeloError as e:
        print(f"Server refused HELO message: {e}")
    except smtplib.SMTPSenderRefused as e:
        print(f"Server refused sender address: {e}")
    except smtplib.SMTPRecipientsRefused as e:
        print(f"Server refused recipient addresses: {e}")
    except smtplib.SMTPDataError as e:
        print(f"Unexpected error code after DATA command: {e}")
    except smtplib.SMTPException as e:
        print(f"SMTP error occurred: {e}")
    except Exception as e:
        print(f"An error occurred: {e}")
         
         
def alert_by_email(body:str, charger_name:str) -> None:
    subject = f'Rectifier level alert @{charger_name}'
    from_email = 'sitthi.s@symphony.net.th'
    #to_email = ['sitthi.s@symphony.net.th']
    to_email = ['sitthi.s@symphony.net.th', 'FacilitiesManagement@symphony.net.th']
    send_email(subject, body, from_email, to_email)

def find_avg_percentage_argus(data:dict) -> str:
    sum = 0
    for i in range(1, len(data['data']['Rectifier'])):
        sum += int(data['data']['Rectifier'][i]['@G'])
    return sum / (len(data['data']['Rectifier'])-1)

def find_avg_percentage_dongah(data:dict) -> str:
    sum = 0
    for i in range(1, len(data['data']['Rectifier'])):
        sum += int(data['data']['Rectifier'][i]['@G'])
    return sum / (len(data['data']['Rectifier'])-1)

def find_avg_percentage_greatpower(html_text:str) -> str:
    soup = BeautifulSoup(html_text, 'html.parser')
    # Find the script tag containing strRectVolt
    script_tags = soup.find_all('script')
    volt_values = []
    for script in script_tags:
        script_text = str(script)
        if 'strRectVolt' in script_text:
            # Extract the array values using string manipulation
            start_idx = script_text.find('strRectVolt=new Array(') + len('strRectVolt=new Array(')
            end_idx = script_text.find(');', start_idx)
            if start_idx > 0 and end_idx > start_idx:
                volt_str = script_text[start_idx:end_idx].split(',')
                for volt in volt_str:
                    if volt == '0.0': continue
                    try:
                        volt_values.append(float(volt))
                    except ValueError:
                        continue
                try:
                    print(volt_values)
                    return sum(volt_values) / len(volt_values) if volt_values else 0
                except ValueError:
                    print(f"Warning: Could not parse voltage values from:")
                    return 0 
    return 0


def check_high_utilization(charger_name: str, percentage: float) -> dict:
    """Check if charger utilization is above 80% and take appropriate action."""
    current_time = datetime.now()
    
    if float(percentage) > 70:
        return {
            'charger_name': charger_name,
            'percentage': percentage,
            'time': current_time
        }
    return None
    

def fetch_data_from_api(charger_name:str, charger_ip:str, charger_model:str):
    url = charger_url[charger_model]['url'].replace('charger_ip',charger_ip)
    user = charger_url[charger_model]['user']
    passwd = charger_url[charger_model]['pass']
    #print(url)
    try:
        response = requests.get(url, auth=(user, passwd), timeout=2)
        response.raise_for_status()
    except requests.exceptions.RequestException as e:
        print(f"API request error: {e}")
        return {charger_name: "Failed to fetch data from API"}
    
    #print(response.text)
    if charger_model == 'Argus':
        percentage = find_avg_percentage_argus(response.text)
    elif charger_model == 'Great power':
        pass
        #percentage = find_avg_percentage_greatpower(response.text)
    elif charger_model == 'DongAH': 
        pass
        #percentage = find_avg_percentage_dongah(response.text)
    print(f"INFO: {charger_name} utilization is {percentage:.2f}% - Normal operation")
    return {charger_name: percentage}
    # Add other charger model handling here
    return {charger_name: "Unsupported charger model"}

def fetch_all_apis(Charger_dict):
    charger_list = [name for name, info in Charger_dict.items() if info['Brand'] == 'Great power' and info['Status'] == 'On-line']
    ip_list = [info['IP'] for info in Charger_dict.values() if info['Brand'] == 'Great power' and info['Status'] == 'On-line']
    model_list = [info['Brand'] for info in Charger_dict.values() if info['Brand'] == 'Great power' and info['Status'] == 'On-line']
    """Fetch data from all API endpoints"""
    pool = ThreadPool(8)
    return pool.starmap(fetch_data_from_api,zip(charger_list, ip_list, model_list))
    
def update_api_data_periodically():
    """Background thread function to update API data every 10 seconds"""
    global api_data_cache, last_update_time
    
    Charger_dict = read_charger_list()

    while True:
        print("Fetching fresh API data...")
        #ip_list = [info['IP'] for info in Charger_dict.values()]
        #ping_result = multithread_ping(ip_list)
        #update_status_charger(ping_result)
        
        api_data_cache = fetch_all_apis(Charger_dict)
        last_update_time = time.strftime("%Y-%m-%d %H:%M:%S")
        
        """ high_utilization_alerts = []
        for result in api_data_cache:
            for charger_name, percentage in result.items():
                if percentage == 'Failed to fetch data from API':continue
                alert_info = check_high_utilization(charger_name, percentage)
                if alert_info:
                    high_utilization_alerts.append(alert_info)
        
        # Send consolidated email if there are any alerts
        if high_utilization_alerts:
            body = "WARNING: High Utilization Alerts:\n\n"
            for alert in high_utilization_alerts:
                body += f"- {alert['charger_name']}: {alert['percentage']:.2f}% (Checked at {alert['time'].strftime('%Y-%m-%d %H:%M:%S')})\n"
            
            try:
                alert_by_email(body, "Multiple Chargers")
                print(f"Alert sent for {len(high_utilization_alerts)} chargers at {last_update_time}")
            except Exception as e:
                print(f"ERROR: Failed to send consolidated alert: {str(e)}") 
        
        with open("Result.txt","w") as result:
            for line in api_data_cache:
                result.write(str(line)+'\n') """
        print('###########################')
        print('\nWaiting for next round...')
        print('###########################')    
        time.sleep(10*60)

@app.route('/')
def index():
    return render_template('index.html', 
                           api_data=api_data_cache, 
                           last_update=last_update_time)

@app.route('/api/data')
def get_api_data():
    """API endpoint to get the latest data via AJAX"""
    return jsonify({
        'api_data': api_data_cache,
        'last_update': last_update_time
    })

if __name__ == "__main__":
    # Initialize the update thread
    update_thread = threading.Thread(target=update_api_data_periodically, daemon=True)
    update_thread.start()
    
    # Run the Flask app with debug mode disabled
    app.run(host='0.0.0.0', debug=False)